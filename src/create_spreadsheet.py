import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "docs/synthesis_calculator.xlsx"

def create_spreadsheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Synthesis Calculator"

    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    input_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid") # Light Green
    formula_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light Grey
    result_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center")

    # --- Constants Section ---
    ws["A1"] = "Constants"
    ws["A1"].font = Font(bold=True, size=14)
    
    constants = [
        ("Cost per Gear ($x)", 7),
        ("Cost per Area ($y)", 0.1),
        ("Cost per Slot ($z)", 10),
        ("Target Pos (H14)", 26)
    ]
    
    for i, (name, val) in enumerate(constants, start=2):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = val
        ws[f"B{i}"].fill = input_fill

    # Named Ranges (simulated by cell references for simplicity in this script)
    # B2 = x, B3 = y, B4 = z, B5 = Target

    # --- Input Table Header ---
    headers = ["Gear ID", "Shaft Pos (cm)", "Radius (cm)", "Area (cm^2)", "Area Cost ($)", "Gear Cost ($)"]
    start_row = 8
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # --- Data Rows (Example: Simple Optimized) ---
    # H1, H4, H7, H10, H14 -> 0, 6, 12, 18, 26
    # Radii: 0.2857, 5.7143, 0.2857, 5.7143, 2.2857
    example_data = [
        (1, 0, 0.2857),
        (2, 6, 5.7143),
        (3, 12, 0.2857),
        (4, 18, 5.7143),
        (5, 26, 2.2857)
    ]

    for i, (gid, pos, rad) in enumerate(example_data, start=start_row+1):
        ws.cell(row=i, column=1, value=gid).alignment = center_align
        
        # Input Columns (Green)
        c_pos = ws.cell(row=i, column=2, value=pos)
        c_pos.fill = input_fill
        c_pos.alignment = center_align
        
        c_rad = ws.cell(row=i, column=3, value=rad)
        c_rad.fill = input_fill
        c_rad.alignment = center_align
        
        # Formulas
        # Area = PI * r^2
        ws.cell(row=i, column=4, value=f"=PI()*C{i}^2").fill = formula_fill
        
        # Area Cost = Area * y ($B$3)
        ws.cell(row=i, column=5, value=f"=D{i}*$B$3").fill = formula_fill
        
        # Gear Cost = IF(Radius>0, x, 0) ($B$2)
        ws.cell(row=i, column=6, value=f"=IF(C{i}>0, $B$2, 0)").fill = formula_fill

    # --- Summary Section ---
    sum_row = start_row + len(example_data) + 2
    
    ws[f"D{sum_row}"] = "Subtotals:"
    ws[f"D{sum_row}"].font = Font(bold=True)
    ws[f"D{sum_row}"].alignment = Alignment(horizontal="right")
    
    # Sum Area Cost
    ws[f"E{sum_row}"] = f"=SUM(E{start_row+1}:E{start_row+len(example_data)})"
    ws[f"E{sum_row}"].font = Font(bold=True)
    
    # Sum Gear Cost
    ws[f"F{sum_row}"] = f"=SUM(F{start_row+1}:F{start_row+len(example_data)})"
    ws[f"F{sum_row}"].font = Font(bold=True)

    # --- Final Calculation ---
    final_row = sum_row + 2
    
    # Slot Calculation
    ws[f"A{final_row}"] = "Output Shaft Pos:"
    ws[f"B{final_row}"] = f"=B{start_row+len(example_data)}" # Last shaft pos
    
    ws[f"A{final_row+1}"] = "Slot Displacement:"
    ws[f"B{final_row+1}"] = f"=B{final_row}-$B$5" # Pos - 26
    
    ws[f"A{final_row+2}"] = "Slot Penalty ($):"
    ws[f"B{final_row+2}"] = f"=$B$4 * ABS(B{final_row+1})^3" # z * |disp|^3
    ws[f"B{final_row+2}"].font = Font(bold=True, color="FF0000")

    # Total Score
    ws[f"D{final_row}"] = "TOTAL SCORE:"
    ws[f"D{final_row}"].font = Font(bold=True, size=14)
    ws[f"D{final_row}"].alignment = Alignment(horizontal="right")
    
    ws[f"E{final_row}"] = f"=E{sum_row} + F{sum_row} + B{final_row+2}"
    ws[f"E{final_row}"].font = Font(bold=True, size=16, color="00B050")
    ws[f"E{final_row}"].fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")

    # --- Column Widths ---
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    wb.save(OUTPUT_FILE)
    print(f"Spreadsheet saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    create_spreadsheet()
